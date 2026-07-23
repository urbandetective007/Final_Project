#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import zipfile
import os

today = "23-07-2026"
filename = f"דוח נכסים לתאריך:{today}.xlsx"
filepath = f"/home/user/Final_Project/excel outputs/{filename}"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ממצאים"
ws.sheet_view.rightToLeft = True

# Headers (row 1)
headers = ["שם העסק", "כתובת", "סוג עסק", "דירוג אינדיקציה", "קישור 1", "קישור 2", "קישור 3"]
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)

for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Business data
businesses = [
    # (שם העסק, כתובת, סוג עסק, קישור1, קישור2, קישור3)
    ("יש חסד", "ברכת אברהם 22, ירושלים", "סופרמרקטים ומכולות",
     "https://www.d.co.il/80056577/35760/", "http://www.yesh.co.il", None),
    ("בית מרקחת- מאוחדת", "ברכת אברהם 16, ירושלים", "בתי מרקחת",
     "https://www.d.co.il/80123798/8150/", "http://www.meuhedet.co.il", None),
    ("מחסן הבשר", "ברכת אברהם 4, ירושלים", "קצביות",
     "https://www.d.co.il/80012629/2490/", None, None),
    ("אורנג' מירה", "ברכת אברהם 16, ירושלים", "הריון ולידה",
     "https://www.d.co.il/80248063/14165/", None, None),
    ("מאוחדת", "ברכת אברהם 16, ירושלים", "קופות חולים",
     "https://www.d.co.il/80270680/43950/", "https://www.meuhedet.co.il/", None),
    ("בן-עמי שוקי", "ברנשטיין פרץ 22, ירושלים", "עורכי דין",
     "https://www.d.co.il/233090/38190/", None, None),
    ("ד\"ר חמו יצחק", "ברנשטיין פרץ 46, ירושלים", "רפואת עיניים",
     "https://d.co.il/24526470/46560", None, None),
    ("אוחנה מוטי", "ברנשטיין פרץ 6, ירושלים", "שירותים",
     "https://www.d.co.il/25707940/14625/", None, None),
    ("סטודיו דוב אברמסון", "ברניקי 7, ירושלים", "עיצוב גרפי",
     "https://www.d.co.il/80184921/11380/", None, None),
    ("סלכל בו - סלסילה", "גבעת שאול 24, ירושלים", "סופרמרקטים ומכולות",
     "https://www.d.co.il/12702890/35760/", None, None),
    ("המלך החסיד", "גבעת שאול 25, ירושלים", "הדפסה דיגיטלית",
     "https://www.d.co.il/76762080/12390/", None, None),
    ("פריד הנדסה בהנהלת פרידבורג חזקיהו", "גבעת שאול 24, ירושלים", "הנדסה ותכנון",
     "https://www.d.co.il/80140237/26270/", None, None),
    ("טליס - tali's", "גבעת קנדה 16, ירושלים", "עיצוב אריזות",
     "https://www.d.co.il/80068899/3970/", None, None),
    ("כהן סימה", "גבעת קנדה 31, ירושלים", "מורים לנהיגה",
     "https://www.d.co.il/37797660/8450/", None, None),
    ("מדיה - ווייז", "גדוד חרמש 16, ירושלים", "מדיה ופרסום",
     "https://www.d.co.il/7351750/40920/", None, None),
    ("דוד שירותי צבע", "גדוד חרמש 16, ירושלים", "צבעים וסיידים",
     "https://www.d.co.il/80087602/41160/", None, None),
    ("אפי נכסים", "גדוד חרמש 8, ירושלים", "מתווכים ומשרדי תיווך",
     "https://www.d.co.il/71534970/32970/", None, None),
    ("אברהם יואב", "גדוד חרמש 15, ירושלים", "חשמלאי",
     "https://www.d.co.il/11809420/17910/", None, None),
    ("דלאל מוטי", "גולומב 23, ירושלים", "שירותי ייעוץ",
     "https://www.d.co.il/76691530/2410/", None, None),
    ("לויתן בני", "גלבר 14, ירושלים", "גישור ויישוב סכסוכים",
     "https://www.d.co.il/71620430/9880/", None, None),
]

green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
blue_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for row_idx, biz in enumerate(businesses, start=2):
    name, addr, biz_type, link1, link2, link3 = biz
    row_fill = green_fill if row_idx % 2 == 0 else blue_fill

    ws.cell(row=row_idx, column=1, value=name).fill = row_fill
    ws.cell(row=row_idx, column=2, value=addr).fill = row_fill
    ws.cell(row=row_idx, column=3, value=biz_type).fill = row_fill
    ws.cell(row=row_idx, column=4, value="גבוה").fill = row_fill
    ws.cell(row=row_idx, column=5, value=link1).fill = row_fill
    ws.cell(row=row_idx, column=6, value=link2).fill = row_fill
    ws.cell(row=row_idx, column=7, value=link3).fill = row_fill

    for col in range(1, 8):
        ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="right", vertical="center")

# Set column widths
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 50
ws.column_dimensions['F'].width = 40
ws.column_dimensions['G'].width = 40

wb.save(filepath)
print(f"Saved: {filepath}")

# Verify it's valid
with zipfile.ZipFile(filepath, 'r') as z:
    assert z.testzip() is None, "Excel file corrupted"
print(f"Valid Excel. Size: {os.path.getsize(filepath)} bytes")
print(f"Total businesses: {len(businesses)}")
